from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_required, current_user
from datetime import datetime, timedelta, date
from models import db, User, WorkEntry, Vacation, Shift, OvertimeRequest
from auth import init_auth_routes
from utils import admin_required
import calendar
from sqlalchemy import func
import os
import openpyxl
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'

# Настройка базы данных: PostgreSQL на сервере, SQLite для локальной разработки
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL + '?sslmode=require'
else:
    basedir = os.path.abspath(os.path.dirname(__file__))
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'database.db')

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

init_auth_routes(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.is_admin():
            return redirect(url_for('admin_calendar'))
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    
    work_entries = WorkEntry.query.filter(
        WorkEntry.user_id == current_user.id,
        WorkEntry.date >= start_of_month,
        WorkEntry.date <= today
    ).all()
    
    total_hours = sum(entry.hours_worked for entry in work_entries)
    
    vacations = Vacation.query.filter(
        Vacation.user_id == current_user.id,
        Vacation.status == 'approved',
        Vacation.start_date <= today,
        Vacation.end_date >= start_of_month
    ).all()
    
    vacation_days = sum((vac.end_date - vac.start_date).days + 1 for vac in vacations)
    
    shifts = Shift.query.filter(
        Shift.user_id == current_user.id,
        Shift.date >= start_of_month,
        Shift.date <= today + timedelta(days=30)
    ).order_by(Shift.date).all()
    
    return render_template('dashboard.html',
                         total_hours=total_hours,
                         vacation_days=vacation_days,
                         work_entries=work_entries[-10:],
                         shifts=shifts)

@app.route('/work_log', methods=['GET', 'POST'])
@login_required
def work_log():
    if request.method == 'POST':
        date_str = request.form.get('date')
        hours_worked = float(request.form.get('hours_worked'))
        description = request.form.get('description')
        
        entry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        existing = WorkEntry.query.filter_by(
            user_id=current_user.id,
            date=entry_date
        ).first()
        
        if existing:
            flash('Запись за эту дату уже существует', 'error')
            return redirect(url_for('work_log'))
        
        work_entry = WorkEntry(
            user_id=current_user.id,
            date=entry_date,
            hours_worked=hours_worked,
            description=description
        )
        
        db.session.add(work_entry)
        db.session.commit()
        flash('Рабочий день успешно добавлен!', 'success')
        return redirect(url_for('work_log'))
    
    entries = WorkEntry.query.filter_by(user_id=current_user.id).order_by(WorkEntry.date.desc()).all()
    return render_template('work_log.html', entries=entries)

@app.route('/vacation_request', methods=['POST'])
@login_required
def vacation_request():
    start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
    end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
    vacation_type = request.form.get('type')
    reason = request.form.get('reason')
    
    existing = Vacation.query.filter_by(
        user_id=current_user.id,
        start_date=start_date,
        status='pending'
    ).first()
    
    if existing:
        flash('У вас уже есть активная заявка на этот период', 'error')
        return redirect(url_for('dashboard'))
    
    vacation = Vacation(
        user_id=current_user.id,
        start_date=start_date,
        end_date=end_date,
        type=vacation_type,
        reason=reason
    )
    
    db.session.add(vacation)
    db.session.commit()
    flash('Заявка на выходные отправлена', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/calendar')
@login_required
@admin_required
def admin_calendar():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    today = date.today()
    if not year or not month:
        year = today.year
        month = today.month
    
    users = User.query.filter(User.is_active == True, User.username != 'admin').all()
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    shifts = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    ).all()
    
    shifts_dict = {}
    for shift in shifts:
        shifts_dict[(shift.user_id, shift.date)] = shift
    
    vacations = Vacation.query.filter(
        Vacation.status == 'approved',
        Vacation.start_date <= end_date,
        Vacation.end_date >= start_date
    ).all()
    
    vacations_dict = {}
    for vac in vacations:
        current = vac.start_date
        while current <= vac.end_date:
            if start_date <= current <= end_date:
                vacations_dict[(vac.user_id, current)] = vac
            current += timedelta(days=1)
    
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    
    days_in_month = calendar.monthrange(year, month)[1]
    days = []
    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        days.append({
            'date': current_date,
            'day_of_week': current_date.weekday(),
            'is_weekend': current_date.weekday() >= 5
        })
    
    return render_template('admin_calendar.html',
                         users=users,
                         days=days,
                         shifts_dict=shifts_dict,
                         vacations_dict=vacations_dict,
                         year=year,
                         month=month,
                         month_name=month_names[month-1],
                         prev_year=prev_year,
                         prev_month=prev_month,
                         next_year=next_year,
                         next_month=next_month)

@app.route('/admin/add_shift_ajax', methods=['POST'])
@login_required
@admin_required
def add_shift_ajax():
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        date_str = data.get('date')
        shift_type = data.get('shift_type')
        
        # Получаем время из запроса
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        # Если время не передано или это стандартное значение - подставляем стандартное для типа смены
        if not start_time or not end_time or (start_time == '09:00' and end_time == '18:00'):
            if shift_type == 'morning':
                start_time = '09:00'
                end_time = '18:00'
            elif shift_type == 'day':
                start_time = '10:00'
                end_time = '19:00'
            elif shift_type == 'night':
                start_time = '16:00'
                end_time = '23:00'
            else:
                start_time = '09:00'
                end_time = '18:00'
        
        shift_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Расчет часов
        start_hour = int(start_time.split(':')[0])
        start_minute = int(start_time.split(':')[1])
        end_hour = int(end_time.split(':')[0])
        end_minute = int(end_time.split(':')[1])
        
        if end_hour < start_hour or (end_hour == start_hour and end_minute < start_minute):
            hours_worked = (24 - start_hour - start_minute/60) + end_hour + end_minute/60
        else:
            hours_worked = (end_hour - start_hour) + (end_minute - start_minute)/60
        
        if shift_type == 'vacation' or shift_type == 'off':
            hours_worked = 0
        elif hours_worked > 4:
            hours_worked = hours_worked - 1
        
        hours_worked = round(hours_worked, 1)
        
        # Сохраняем смену
        existing_shift = Shift.query.filter_by(user_id=user_id, date=shift_date).first()
        
        if existing_shift:
            existing_shift.shift_type = shift_type
            existing_shift.start_time = start_time
            existing_shift.end_time = end_time
        else:
            shift = Shift(
                user_id=user_id,
                date=shift_date,
                shift_type=shift_type,
                start_time=start_time,
                end_time=end_time
            )
            db.session.add(shift)
        
        # Сохраняем запись о часах
        existing_work = WorkEntry.query.filter_by(user_id=user_id, date=shift_date).first()
        
        description = f"Смена: {shift_type} {start_time}-{end_time}"
        
        if existing_work:
            existing_work.hours_worked = hours_worked
            existing_work.description = description
        else:
            work_entry = WorkEntry(
                user_id=user_id,
                date=shift_date,
                hours_worked=hours_worked,
                description=description
            )
            db.session.add(work_entry)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Смена добавлена', 'start_time': start_time, 'end_time': end_time})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete_shift_ajax', methods=['POST'])
@login_required
@admin_required
def delete_shift_ajax():
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        date_str = data.get('date')
        
        shift_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        shift = Shift.query.filter_by(user_id=user_id, date=shift_date).first()
        if shift:
            db.session.delete(shift)
        
        work_entry = WorkEntry.query.filter_by(user_id=user_id, date=shift_date).first()
        if work_entry:
            db.session.delete(work_entry)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Смена удалена'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin')
@login_required
@admin_required
def admin():
    users = User.query.filter(User.username != 'admin').all()
    pending_vacations = Vacation.query.filter_by(status='pending').all()
    work_entries = WorkEntry.query.filter_by(status='active').all()
    
    return render_template('admin.html', 
                         users=users, 
                         pending_vacations=pending_vacations,
                         work_entries=work_entries)

@app.route('/admin/add_user', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')
    
    working_hours_str = request.form.get('working_hours', '8.0')
    if not working_hours_str or working_hours_str.strip() == '':
        working_hours = 8.0
    else:
        try:
            working_hours = float(working_hours_str)
        except ValueError:
            working_hours = 8.0
    
    if User.query.filter_by(username=username).first():
        flash('Пользователь с таким именем уже существует', 'error')
        return redirect(url_for('admin'))
    
    if User.query.filter_by(email=email).first():
        flash('Пользователь с таким email уже существует', 'error')
        return redirect(url_for('admin'))
    
    user = User(
        username=username,
        email=email,
        role=role,
        working_hours_per_day=working_hours
    )
    user.set_password(password)
    
    db.session.add(user)
    db.session.commit()
    flash('Пользователь успешно добавлен', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/toggle_user/<int:user_id>')
@login_required
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id != current_user.id:
        user.is_active = not user.is_active
        db.session.commit()
        flash(f'Статус пользователя {user.username} изменен', 'success')
    else:
        flash('Нельзя изменить статус самого себя', 'error')
    return redirect(url_for('admin'))

@app.route('/admin/approve_vacation/<int:vac_id>')
@login_required
@admin_required
def approve_vacation(vac_id):
    vacation = Vacation.query.get_or_404(vac_id)
    vacation.status = 'approved'
    db.session.commit()
    flash('Выходные одобрены', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/reject_vacation/<int:vac_id>')
@login_required
@admin_required
def reject_vacation(vac_id):
    vacation = Vacation.query.get_or_404(vac_id)
    vacation.status = 'rejected'
    db.session.commit()
    flash('Выходные отклонены', 'success')
    return redirect(url_for('admin'))

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html')

@app.route('/admin/delete_user/<int:user_id>')
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id != current_user.id:
        WorkEntry.query.filter_by(user_id=user.id).delete()
        Vacation.query.filter_by(user_id=user.id).delete()
        Shift.query.filter_by(user_id=user.id).delete()
        db.session.delete(user)
        db.session.commit()
        flash(f'Пользователь {user.username} удален', 'success')
    else:
        flash('Нельзя удалить самого себя', 'error')
    return redirect(url_for('admin'))

@app.route('/my_shifts')
@login_required
def my_shifts():
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    if today.month == 12:
        end_of_month = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        end_of_month = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    shifts = Shift.query.filter(
        Shift.user_id == current_user.id,
        Shift.date >= start_of_month,
        Shift.date <= end_of_month
    ).order_by(Shift.date).all()
    
    return render_template('my_shifts.html', shifts=shifts)

@app.route('/admin/reports')
@login_required
@admin_required
def admin_reports():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    today = date.today()
    if not year or not month:
        year = today.year
        month = today.month
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    users = User.query.filter(User.is_active == True, User.username != 'admin').all()
    
    user_stats = []
    for user in users:
        shifts = Shift.query.filter(
            Shift.user_id == user.id,
            Shift.date >= start_date,
            Shift.date <= end_date,
            Shift.shift_type.in_(['morning', 'day', 'night'])
        ).all()
        
        days_off = Shift.query.filter(
            Shift.user_id == user.id,
            Shift.date >= start_date,
            Shift.date <= end_date,
            Shift.shift_type == 'off'
        ).all()
        
        vacations = Shift.query.filter(
            Shift.user_id == user.id,
            Shift.date >= start_date,
            Shift.date <= end_date,
            Shift.shift_type == 'vacation'
        ).all()
        
        vacation_requests = Vacation.query.filter(
            Vacation.user_id == user.id,
            Vacation.status == 'approved',
            Vacation.start_date <= end_date,
            Vacation.end_date >= start_date
        ).all()
        
        vacation_days_from_requests = 0
        for vac in vacation_requests:
            vac_start = max(vac.start_date, start_date)
            vac_end = min(vac.end_date, end_date)
            vacation_days_from_requests += (vac_end - vac_start).days + 1
        
        total_vacation_days = len(vacations) + vacation_days_from_requests
        
        work_entries = WorkEntry.query.filter(
            WorkEntry.user_id == user.id,
            WorkEntry.date >= start_date,
            WorkEntry.date <= end_date
        ).all()
        
        shift_count = {
            'morning': 0,
            'day': 0,
            'night': 0,
            'off': len(days_off),
            'vacation': total_vacation_days
        }
        for shift in shifts:
            shift_count[shift.shift_type] += 1
        
        total_hours = sum(entry.hours_worked for entry in work_entries)
        
        user_stats.append({
            'user': user,
            'shifts': shifts,
            'days_off': days_off,
            'vacations': vacations,
            'vacation_days': total_vacation_days,
            'shift_count': shift_count,
            'total_shifts': len(shifts),
            'total_days_off': len(days_off),
            'total_hours': total_hours,
            'work_entries': work_entries
        })
    
    total_shifts_all = sum(stat['total_shifts'] for stat in user_stats)
    total_hours_all = sum(stat['total_hours'] for stat in user_stats)
    total_vacation_days_all = sum(stat['vacation_days'] for stat in user_stats)
    total_days_off_all = sum(stat['total_days_off'] for stat in user_stats)
    
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    
    return render_template('admin_reports.html',
                         user_stats=user_stats,
                         year=year,
                         month=month,
                         month_name=month_names[month-1],
                         prev_year=prev_year,
                         prev_month=prev_month,
                         next_year=next_year,
                         next_month=next_month,
                         total_shifts_all=total_shifts_all,
                         total_hours_all=total_hours_all,
                         total_vacation_days_all=total_vacation_days_all,
                         total_days_off_all=total_days_off_all)

# Запрос на дополнительные часы
@app.route('/overtime_request', methods=['GET', 'POST'])
@login_required
def overtime_request():
    if request.method == 'POST':
        date_str = request.form.get('date')
        hours = float(request.form.get('hours'))
        reason = request.form.get('reason')
        
        request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        existing = OvertimeRequest.query.filter_by(
            user_id=current_user.id,
            date=request_date,
            status='pending'
        ).first()
        
        if existing:
            flash('У вас уже есть активный запрос на эту дату', 'error')
            return redirect(url_for('overtime_request'))
        
        overtime_req = OvertimeRequest(
            user_id=current_user.id,
            date=request_date,
            hours=hours,
            reason=reason
        )
        
        db.session.add(overtime_req)
        db.session.commit()
        flash('Запрос на дополнительные часы отправлен на согласование', 'success')
        return redirect(url_for('dashboard'))
    
    requests = OvertimeRequest.query.filter_by(user_id=current_user.id).order_by(OvertimeRequest.date.desc()).all()
    return render_template('overtime_request.html', requests=requests)

# Админ: просмотр всех запросов
@app.route('/admin/overtime_requests')
@login_required
@admin_required
def admin_overtime_requests():
    pending_requests = OvertimeRequest.query.filter_by(status='pending').order_by(OvertimeRequest.date).all()
    approved_requests = OvertimeRequest.query.filter_by(status='approved').order_by(OvertimeRequest.date.desc()).limit(50).all()
    rejected_requests = OvertimeRequest.query.filter_by(status='rejected').order_by(OvertimeRequest.date.desc()).limit(50).all()
    
    return render_template('admin_overtime_requests.html',
                         pending_requests=pending_requests,
                         approved_requests=approved_requests,
                         rejected_requests=rejected_requests)

# Админ: одобрить запрос
@app.route('/admin/approve_overtime/<int:req_id>')
@login_required
@admin_required
def approve_overtime(req_id):
    overtime_req = OvertimeRequest.query.get_or_404(req_id)
    overtime_req.status = 'approved'
    
    existing_work = WorkEntry.query.filter_by(
        user_id=overtime_req.user_id,
        date=overtime_req.date
    ).first()
    
    if existing_work:
        existing_work.hours_worked += overtime_req.hours
        existing_work.description += f" | +{overtime_req.hours}ч (запрос: {overtime_req.reason})"
    else:
        work_entry = WorkEntry(
            user_id=overtime_req.user_id,
            date=overtime_req.date,
            hours_worked=overtime_req.hours,
            description=f"Запрос на доп. часы: {overtime_req.reason}"
        )
        db.session.add(work_entry)
    
    db.session.commit()
    flash(f'Запрос на {overtime_req.hours} часов одобрен', 'success')
    return redirect(url_for('admin_overtime_requests'))

# Админ: отклонить запрос
@app.route('/admin/reject_overtime/<int:req_id>', methods=['POST'])
@login_required
@admin_required
def reject_overtime(req_id):
    overtime_req = OvertimeRequest.query.get_or_404(req_id)
    overtime_req.status = 'rejected'
    admin_comment = request.form.get('admin_comment', '')
    overtime_req.admin_comment = admin_comment
    db.session.commit()
    flash(f'Запрос на {overtime_req.hours} часов отклонен', 'warning')
    return redirect(url_for('admin_overtime_requests'))

# Админ: добавить часы сотруднику
@app.route('/admin/add_hours/<int:user_id>')
@login_required
@admin_required
def admin_add_hours(user_id):
    user = User.query.get_or_404(user_id)
    return render_template('admin_add_hours.html', user=user)

@app.route('/admin/save_hours', methods=['POST'])
@login_required
@admin_required
def admin_save_hours():
    user_id = request.form.get('user_id', type=int)
    date_str = request.form.get('date')
    hours = float(request.form.get('hours'))
    reason = request.form.get('reason')
    
    work_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    existing_work = WorkEntry.query.filter_by(user_id=user_id, date=work_date).first()
    
    if existing_work:
        existing_work.hours_worked += hours
        existing_work.description += f" | Админ {current_user.username} добавил {hours}ч: {reason}"
    else:
        work_entry = WorkEntry(
            user_id=user_id,
            date=work_date,
            hours_worked=hours,
            description=f"Админ {current_user.username} добавил: {reason}"
        )
        db.session.add(work_entry)
    
    db.session.commit()
    flash(f'Добавлено {hours} часов пользователю', 'success')
    return redirect(url_for('admin'))

# Админ: добавить часы себе
@app.route('/admin/add_my_hours', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_my_hours():
    if request.method == 'POST':
        date_str = request.form.get('date')
        hours = float(request.form.get('hours'))
        reason = request.form.get('reason')
        
        work_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        existing_work = WorkEntry.query.filter_by(user_id=current_user.id, date=work_date).first()
        
        if existing_work:
            existing_work.hours_worked += hours
            existing_work.description += f" | Добавил себе {hours}ч: {reason}"
        else:
            work_entry = WorkEntry(
                user_id=current_user.id,
                date=work_date,
                hours_worked=hours,
                description=f"Добавил себе: {reason}"
            )
            db.session.add(work_entry)
        
        db.session.commit()
        flash(f'Добавлено {hours} часов', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('admin_add_my_hours.html')

# Импорт графика из Excel
@app.route('/admin/import_schedule', methods=['GET', 'POST'])
@login_required
@admin_required
def import_schedule():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не выбран', 'error')
            return redirect(url_for('import_schedule'))
        
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран', 'error')
            return redirect(url_for('import_schedule'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Поддерживаются только файлы Excel (.xlsx, .xls)', 'error')
            return redirect(url_for('import_schedule'))
        
        try:
            workbook = openpyxl.load_workbook(BytesIO(file.read()))
            sheet = workbook.active
            
            users = {user.username: user for user in User.query.filter(User.username != 'admin').all()}
            
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            user_col = 1
            
            date_columns = []
            for col_idx, header in enumerate(headers, start=1):
                if col_idx != user_col:
                    try:
                        date_col = datetime.strptime(str(header), '%Y-%m-%d').date()
                        date_columns.append((col_idx, date_col))
                    except ValueError:
                        pass
            
            if not date_columns:
                flash('Не найдены колонки с датами. Используйте формат ГГГГ-ММ-ДД в заголовках', 'error')
                return redirect(url_for('import_schedule'))
            
            created_count = 0
            updated_count = 0
            errors = []
            
            for row in range(2, sheet.max_row + 1):
                username = sheet.cell(row=row, column=user_col).value
                if not username:
                    continue
                
                username = str(username).strip()
                if username not in users:
                    errors.append(f'Сотрудник "{username}" не найден в системе')
                    continue
                
                user = users[username]
                
                for col_idx, shift_date in date_columns:
                    shift_value = sheet.cell(row=row, column=col_idx).value
                    if not shift_value:
                        continue
                    
                    shift_str = str(shift_value).strip().lower()
                    
                    if shift_str in ['утро', 'morning', 'м', 'am']:
                        shift_type = 'morning'
                        start_time = '09:00'
                        end_time = '18:00'
                    elif shift_str in ['день', 'day', 'д', 'd']:
                        shift_type = 'day'
                        start_time = '10:00'
                        end_time = '19:00'
                    elif shift_str in ['вечер', 'evening', 'night', 'в', 'н', 'pm']:
                        shift_type = 'night'
                        start_time = '16:00'
                        end_time = '23:00'
                    elif shift_str in ['выходной', 'off', 'вых', 'o']:
                        shift_type = 'off'
                        start_time = '00:00'
                        end_time = '00:00'
                    elif shift_str in ['отпуск', 'vacation', 'отп', 'v']:
                        shift_type = 'vacation'
                        start_time = '00:00'
                        end_time = '00:00'
                    else:
                        continue
                    
                    existing_shift = Shift.query.filter_by(user_id=user.id, date=shift_date).first()
                    
                    if existing_shift:
                        existing_shift.shift_type = shift_type
                        existing_shift.start_time = start_time
                        existing_shift.end_time = end_time
                        updated_count += 1
                    else:
                        shift = Shift(
                            user_id=user.id,
                            date=shift_date,
                            shift_type=shift_type,
                            start_time=start_time,
                            end_time=end_time
                        )
                        db.session.add(shift)
                        created_count += 1
                    
                    if shift_type in ['off', 'vacation']:
                        hours_worked = 0
                    else:
                        start_hour = int(start_time.split(':')[0])
                        end_hour = int(end_time.split(':')[0])
                        hours_worked = end_hour - start_hour
                        if hours_worked > 4:
                            hours_worked = hours_worked - 1
                    
                    existing_work = WorkEntry.query.filter_by(user_id=user.id, date=shift_date).first()
                    
                    if existing_work:
                        existing_work.hours_worked = hours_worked
                        existing_work.description = f"Импорт: {shift_type} {start_time}-{end_time}"
                    else:
                        work_entry = WorkEntry(
                            user_id=user.id,
                            date=shift_date,
                            hours_worked=hours_worked,
                            description=f"Импорт: {shift_type} {start_time}-{end_time}"
                        )
                        db.session.add(work_entry)
            
            db.session.commit()
            
            flash(f'Импорт завершён! Создано: {created_count}, Обновлено: {updated_count}', 'success')
            if errors:
                flash(f'Ошибки: {", ".join(errors[:5])}', 'warning')
            
            return redirect(url_for('admin_calendar'))
            
        except Exception as e:
            flash(f'Ошибка при импорте: {str(e)}', 'error')
            return redirect(url_for('import_schedule'))
    
    return render_template('import_schedule.html', now=date.today())

# Экспорт графика в Excel
@app.route('/admin/export_schedule')
@login_required
@admin_required
def export_schedule():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    today = date.today()
    if not year or not month:
        year = today.year
        month = today.month
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    users = User.query.filter(User.is_active == True, User.username != 'admin').all()
    
    shifts = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    ).all()
    
    work_entries = WorkEntry.query.filter(
        WorkEntry.date >= start_date,
        WorkEntry.date <= end_date
    ).all()
    
    # Создаём словари для быстрого доступа
    shifts_dict = {}
    for shift in shifts:
        shifts_dict[(shift.user_id, shift.date)] = shift
    
    # Считаем часы для каждого пользователя
    user_hours = {}
    for user in users:
        total_hours = 0
        for entry in work_entries:
            if entry.user_id == user.id:
                total_hours += entry.hours_worked
        user_hours[user.id] = round(total_hours, 1)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"График_{month}_{year}"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    weekend_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
    total_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Заголовок "Сотрудник"
    ws.cell(row=1, column=1, value="Сотрудник")
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border
    
    # Создаем список дней месяца
    current_date = start_date
    col = 2
    days_in_month = []
    while current_date <= end_date:
        ws.cell(row=1, column=col, value=current_date.strftime('%Y-%m-%d'))
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).alignment = center_align
        ws.cell(row=1, column=col).border = thin_border
        
        days_in_month.append({
            'date': current_date,
            'col': col,
            'is_weekend': current_date.weekday() >= 5
        })
        
        current_date += timedelta(days=1)
        col += 1
    
    # Добавляем колонку для итогов
    total_col = len(days_in_month) + 2
    ws.cell(row=1, column=total_col, value="ИТОГО часов")
    ws.cell(row=1, column=total_col).font = header_font
    ws.cell(row=1, column=total_col).fill = header_fill
    ws.cell(row=1, column=total_col).alignment = center_align
    ws.cell(row=1, column=total_col).border = thin_border
    
    # Заполняем данные сотрудников
    row = 2
    for user in users:
        ws.cell(row=row, column=1, value=user.username)
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        # Заполняем смены по дням
        for day_info in days_in_month:
            shift = shifts_dict.get((user.id, day_info['date']))
            col = day_info['col']
            
            if shift:
                if shift.shift_type == 'off':
                    value = 'выходной'
                elif shift.shift_type == 'vacation':
                    value = 'отпуск'
                else:
                    # Для рабочих смен выводим фактическое время
                    value = f'{shift.start_time}-{shift.end_time}'
            else:
                value = ''
           
            
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).alignment = center_align
            ws.cell(row=row, column=col).border = thin_border
            
            if day_info['is_weekend']:
                ws.cell(row=row, column=col).fill = weekend_fill
        
        # Добавляем итоговые часы для сотрудника
        total_hours = user_hours.get(user.id, 0)
        ws.cell(row=row, column=total_col, value=total_hours)
        ws.cell(row=row, column=total_col).font = Font(bold=True)
        ws.cell(row=row, column=total_col).fill = total_fill
        ws.cell(row=row, column=total_col).alignment = center_align
        ws.cell(row=row, column=total_col).border = thin_border
        
        row += 1
    
    # Автоматическая ширина колонок
    for col_num in range(1, total_col + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'graphic_{year}_{month}.xlsx'
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                email='admin@example.com',
                role='admin',
                working_hours_per_day=8.0
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("=" * 50)
            print("Admin user created successfully!")
            print("Username: admin")
            print("Password: admin123")
            print("=" * 50)
    print("\n" + "=" * 50)
    print("Starting Time Tracker Application...")
    print("Access at: http://localhost:5001")
    print("=" * 50 + "\n")
    app.run(debug=True, host='0.0.0.0', port=8080)